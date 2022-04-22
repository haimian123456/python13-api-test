from openpyxl import load_workbook
class Cases:
    '''测试数据类'''
    def __init__(self):
        self.id=None
        self.module = None
        self.description = None
        self.method=None
        self.url=None
        self.param = None
        self.expectedresult = None
        self.sql=None



class Do_Excel:
    '''excel读写类'''
    def __init__(self,filename):
        self.filename=filename #文件名
        try:
            self.workbook=load_workbook(self.filename)#打开测试文件，初始化后调用方便，不用多次打开文件
        except Exception as e:
            print('文件路径错误:',e)

    #excel读取方法
    def read(self,sheetname):
        try:
            self.sheetname=sheetname #sheet名
            sheet=self.workbook[self.sheetname]#获取sheet
            row_max=sheet.max_row#获取最大行
            cases_list=[]#存储测试数据
            for row in range(2,row_max+1):#循环访问每一行测试数据
                cases = Cases()  # 实例化测试数据类
                cases.id=sheet.cell(row=row,column=1).value#获取测试数据并赋值给测试数据类中的属性
                cases.module = sheet.cell(row=row, column=2).value
                cases.description = sheet.cell(row=row, column=3).value
                cases.method = sheet.cell(row=row, column=4).value
                cases.url = sheet.cell(row=row, column=5).value
                cases.param = sheet.cell(row=row, column=6).value
                cases.expectedresult = sheet.cell(row=row, column=7).value
                cases.sql = sheet.cell(row=row, column=10).value
                cases_list.append(cases)#将测试数据对象到列表中

            return cases_list #返回测试数据列表
        except Exception as e:
            print('未成功打开文件',e)
    #excel写入方法
    def write(self,row,actual,result):
        sheet = self.workbook[self.sheetname]  # 获取sheet
        sheet.cell(row=row,column=8).value=actual#将测试实际结果写回excel中
        sheet.cell(row=row, column=9).value = result  # 将测试通过结果写回excel中
        self.workbook.save(self.filename)#关闭文件

if __name__ == '__main__':
    from common.contants import *

    do=Do_Excel(cases_file).read('login')
    print(do[0].id)